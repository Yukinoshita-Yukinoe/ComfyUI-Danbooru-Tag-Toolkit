#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from shutil import copy2
from typing import Iterable


DEFAULT_SOURCE_GLOB = "*.xlsx"
DEFAULT_TEMPLATE_NAME = "translation_map_i18n.json"
DEFAULT_OUTPUT_SUFFIX = "_i18n"
SUPPORTED_DIMENSIONS = ("category", "subcategory")
LEGACY_COLUMNS = {
    "category": ("category", "category_key", "category_zh", "category_en"),
    "subcategory": ("subcategory", "subcategory_key", "subcategory_zh", "subcategory_en"),
}


class TranslationWorkbookError(RuntimeError):
    pass


def require_openpyxl():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise TranslationWorkbookError(
            "Missing dependency 'openpyxl'. Install project requirements first: pip install -r requirements.txt"
        ) from exc
    return load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a translation template and build single-workbook i18n copies with *_key / *_zh / *_en columns."
    )
    parser.add_argument("command", choices=("extract", "migrate"))
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "tags_database",
        help="Directory containing source workbooks. Defaults to tags_database next to this script.",
    )
    parser.add_argument(
        "--glob",
        default=DEFAULT_SOURCE_GLOB,
        help=f"Workbook glob pattern inside input-dir. Default: {DEFAULT_SOURCE_GLOB}",
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=None,
        help=f"Path to the translation template JSON. Default: <input-dir>/{DEFAULT_TEMPLATE_NAME}",
    )
    parser.add_argument(
        "--suffix",
        default=DEFAULT_OUTPUT_SUFFIX,
        help=f"Output suffix for migrated workbook copies. Default: {DEFAULT_OUTPUT_SUFFIX}",
    )
    parser.add_argument(
        "--copy-if-exists",
        action="store_true",
        help="Re-copy from source before writing if the migrated target already exists.",
    )
    return parser.parse_args()


def resolve_template_path(input_dir: Path, template_arg: Path | None) -> Path:
    if template_arg is not None:
        return template_arg.resolve()
    return (input_dir / DEFAULT_TEMPLATE_NAME).resolve()


def discover_workbooks(input_dir: Path, pattern: str, migrated_suffix: str) -> list[Path]:
    if not input_dir.exists():
        raise TranslationWorkbookError(f"Input directory does not exist: {input_dir}")
    workbooks = sorted(
        path
        for path in input_dir.glob(pattern)
        if path.is_file()
        and not path.name.startswith("~$")
        and not path.stem.endswith(str(migrated_suffix or "").strip())
    )
    if not workbooks:
        raise TranslationWorkbookError(f"No workbooks matched '{pattern}' in {input_dir}")
    return workbooks


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slugify(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z]+", "_", value).strip("_").lower()
    return re.sub(r"_+", "_", cleaned)


def next_default_key(prefix: str, source_text: str, used_keys: set[str], index: int) -> str:
    slug = slugify(source_text)
    if slug:
        candidate = f"{prefix}_{slug}"
    else:
        candidate = f"{prefix}_{index:03d}"
    base = candidate
    suffix = 2
    while candidate in used_keys:
        candidate = f"{base}_{suffix}"
        suffix += 1
    used_keys.add(candidate)
    return candidate


def load_template(template_path: Path) -> dict[str, object]:
    if not template_path.exists():
        return {
            "meta": {},
            "category": {},
            "subcategory": {},
        }
    with template_path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise TranslationWorkbookError(f"Template must be a JSON object: {template_path}")
    for dimension in SUPPORTED_DIMENSIONS:
        value = data.get(dimension, {})
        if not isinstance(value, dict):
            raise TranslationWorkbookError(f"'{dimension}' must be a JSON object in {template_path}")
    return data


def read_unique_terms(workbook_path: Path) -> dict[str, list[str]]:
    load_workbook = require_openpyxl()
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    collected = {dimension: set() for dimension in SUPPORTED_DIMENSIONS}
    try:
        for worksheet in workbook.worksheets:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue
            header_map = {
                normalize_text(name): index
                for index, name in enumerate(header_row)
                if normalize_text(name)
            }
            for dimension in SUPPORTED_DIMENSIONS:
                if dimension not in header_map:
                    continue
                column_index = header_map[dimension]
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if column_index >= len(row):
                        continue
                    value = normalize_text(row[column_index])
                    if value:
                        collected[dimension].add(value)
    finally:
        workbook.close()
    return {
        dimension: sorted(values)
        for dimension, values in collected.items()
    }


def write_template(template_path: Path, workbooks: list[Path]) -> None:
    existing = load_template(template_path)
    used_keys = set()
    merged: dict[str, dict[str, dict[str, str]]] = {
        "category": {},
        "subcategory": {},
    }

    for dimension in SUPPORTED_DIMENSIONS:
        existing_entries = existing.get(dimension, {})
        for source_text, payload in existing_entries.items():
            source = normalize_text(source_text)
            if not source or not isinstance(payload, dict):
                continue
            key = normalize_text(payload.get("key"))
            zh = normalize_text(payload.get("zh")) or source
            en = normalize_text(payload.get("en"))
            if key:
                used_keys.add(key)
            merged[dimension][source] = {
                "key": key,
                "zh": zh,
                "en": en,
            }

    dimension_index = {dimension: 1 for dimension in SUPPORTED_DIMENSIONS}
    for workbook_path in workbooks:
        scan = read_unique_terms(workbook_path)
        for dimension, values in scan.items():
            prefix = "cat" if dimension == "category" else "subcat"
            for source_text in values:
                entry = merged[dimension].setdefault(
                    source_text,
                    {"key": "", "zh": source_text, "en": ""},
                )
                if not entry["key"]:
                    entry["key"] = next_default_key(prefix, entry["en"] or source_text, used_keys, dimension_index[dimension])
                    dimension_index[dimension] += 1

    payload = {
        "meta": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source_files": [path.name for path in workbooks],
            "instructions": [
                "Edit key/zh/en values for category and subcategory entries.",
                "Keep keys stable after workflows start using them.",
                "Then run: python scripts/translate_tag_workbooks.py migrate",
            ],
        },
        "category": merged["category"],
        "subcategory": merged["subcategory"],
    }
    template_path.parent.mkdir(parents=True, exist_ok=True)
    with template_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def build_value_lookup(template_data: dict[str, object], dimension: str) -> dict[str, dict[str, str]]:
    raw_entries = template_data.get(dimension, {})
    lookup: dict[str, dict[str, str]] = {}
    if not isinstance(raw_entries, dict):
        return lookup
    for source_text, payload in raw_entries.items():
        source = normalize_text(source_text)
        if not source or not isinstance(payload, dict):
            continue
        key = normalize_text(payload.get("key"))
        zh = normalize_text(payload.get("zh")) or source
        en = normalize_text(payload.get("en")) or zh or source
        if not key:
            raise TranslationWorkbookError(f"Missing '{dimension}.\"{source}\".key' in translation template.")
        lookup[source] = {
            "key": key,
            "zh": zh,
            "en": en,
        }
    return lookup


def migrate_workbook(
    workbook_path: Path,
    template_data: dict[str, object],
    suffix: str,
    copy_if_exists: bool,
) -> tuple[Path, int]:
    load_workbook = require_openpyxl()
    output_path = workbook_path.with_name(f"{workbook_path.stem}{suffix}{workbook_path.suffix}")
    if copy_if_exists or not output_path.exists():
        copy2(workbook_path, output_path)

    category_lookup = build_value_lookup(template_data, "category")
    subcategory_lookup = build_value_lookup(template_data, "subcategory")
    workbook = load_workbook(output_path)
    changed_cells = 0

    try:
        for worksheet in workbook.worksheets:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1), None)
            if not header_row:
                continue

            header_map = {
                normalize_text(cell.value): cell.column
                for cell in header_row
                if normalize_text(cell.value)
            }

            required_missing = [name for name in ("category", "subcategory") if name not in header_map]
            if required_missing:
                continue

            next_column = len(header_row) + 1
            for dimension in SUPPORTED_DIMENSIONS:
                _, key_header, zh_header, en_header = LEGACY_COLUMNS[dimension]
                for header_name in (key_header, zh_header, en_header):
                    if header_name not in header_map:
                        worksheet.cell(row=1, column=next_column, value=header_name)
                        header_map[header_name] = next_column
                        next_column += 1

            for row_index in range(2, worksheet.max_row + 1):
                category_source = normalize_text(worksheet.cell(row=row_index, column=header_map["category"]).value)
                subcategory_source = normalize_text(worksheet.cell(row=row_index, column=header_map["subcategory"]).value)

                if category_source:
                    payload = category_lookup.get(category_source)
                    if not payload:
                        raise TranslationWorkbookError(
                            f"Missing category translation for '{category_source}' while migrating {workbook_path.name}"
                        )
                    updates = {
                        "category_key": payload["key"],
                        "category_zh": payload["zh"],
                        "category_en": payload["en"],
                    }
                    for header_name, value in updates.items():
                        cell = worksheet.cell(row=row_index, column=header_map[header_name])
                        if normalize_text(cell.value) != value:
                            cell.value = value
                            changed_cells += 1

                if subcategory_source:
                    payload = subcategory_lookup.get(subcategory_source)
                    if not payload:
                        raise TranslationWorkbookError(
                            f"Missing subcategory translation for '{subcategory_source}' while migrating {workbook_path.name}"
                        )
                    updates = {
                        "subcategory_key": payload["key"],
                        "subcategory_zh": payload["zh"],
                        "subcategory_en": payload["en"],
                    }
                    for header_name, value in updates.items():
                        cell = worksheet.cell(row=row_index, column=header_map[header_name])
                        if normalize_text(cell.value) != value:
                            cell.value = value
                            changed_cells += 1
        workbook.save(output_path)
    finally:
        workbook.close()

    return output_path, changed_cells


def run_extract(args: argparse.Namespace) -> None:
    workbooks = discover_workbooks(args.input_dir.resolve(), args.glob, args.suffix)
    template_path = resolve_template_path(args.input_dir.resolve(), args.template)
    write_template(template_path, workbooks)
    print(f"[extract] Template written: {template_path}")


def run_migrate(args: argparse.Namespace) -> None:
    workbooks = discover_workbooks(args.input_dir.resolve(), args.glob, args.suffix)
    template_path = resolve_template_path(args.input_dir.resolve(), args.template)
    template_data = load_template(template_path)

    total_changed = 0
    for workbook_path in workbooks:
        output_path, changed_cells = migrate_workbook(
            workbook_path=workbook_path,
            template_data=template_data,
            suffix=args.suffix,
            copy_if_exists=args.copy_if_exists,
        )
        total_changed += changed_cells
        print(f"[migrate] Wrote: {output_path} (changed cells: {changed_cells})")

    print(f"[migrate] Completed. Total changed cells: {total_changed}")


def main() -> int:
    args = parse_args()
    try:
        if args.command == "extract":
            run_extract(args)
        else:
            run_migrate(args)
    except TranslationWorkbookError as exc:
        print(f"Error: {exc}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
